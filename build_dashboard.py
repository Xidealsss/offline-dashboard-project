#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 16 字段 Excel 模板生成自包含的离线项目看板 HTML。

前端源码位于 assets/：template.html / dashboard.css / dashboard.js。
构建时把三者内联成单个自包含 HTML，产物不引用任何外部资源。
开发时可用 --dev 生成 assets/data.js，然后直接打开 assets/template.html 调试。

用法:
    python3 build_dashboard.py [--input 家电产品项目_任务数据.xlsx]
                               [--output 项目看板.html]
                               [--emit-json] [--dev] [--traceback]

退出码: 0 成功 / 1 数据校验失败 / 2 用法或读写错误
"""

import argparse
import json
import re
import sys
import traceback as _traceback
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel

warnings.filterwarnings("ignore")

ASSET_DIR = Path(__file__).resolve().parent / "assets"
ASSET_NAMES = ("template.html", "dashboard.css", "dashboard.js")

# assets/template.html 里的三个外链标签，构建时被替换成内联块
TAG_CSS = '<link rel="stylesheet" href="dashboard.css">'
TAG_DATA = '<script src="data.js"></script>'
TAG_JS = '<script src="dashboard.js"></script>'

HEADERS = [
    "任务ID",
    "阶段",
    "任务",
    "子任务",
    "负责人",
    "协作部门",
    "计划开始",
    "计划结束",
    "进度%",
    "前置任务",
    "里程碑",
    "优先级",
    "状态",
    "风险等级",
    "交付物",
    "备注",
]

VALID_MILESTONE = {"是", "否"}
VALID_PRIORITY = {"高", "中", "低"}
VALID_STATUS = {"未开始", "进行中", "待评审", "已完成", "阻塞"}
VALID_RISK = {"高", "中", "低"}

DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日")

# 输入护栏：一个年份手误（把 2026 打成 2226）就能撑出 7 万天的时间轴，
# 表头要建 14 万个 DOM 节点、泳道宽 200 万像素，浏览器直接卡死。
MAX_SPAN_DAYS = 1095        # 3 年
MAX_ROWS = 800

# 「选项字典」页里参与校验的字段 → 任务字典里的取值方式
DICT_FIELDS = {
    "阶段": lambda t: [t["stage"]],
    "负责人": lambda t: t["owners"],
    "协作部门": lambda t: [t["dept"]] if t["dept"] else [],
    "优先级": lambda t: [t["priority"]],
    "状态": lambda t: [t["status"]],
    "风险等级": lambda t: [t["risk"]],
}


# ---------------------------------------------------------------- 问题收集

@dataclass
class Issue:
    """一条校验问题。库函数只负责收集，由 main() 统一输出并决定退出码。"""

    message: str
    where: str = ""
    level: str = "error"          # error | warning
    hint: str = ""

    def format(self):
        tag = "[校验失败]" if self.level == "error" else "[提示]"
        body = f"{self.where}：{self.message}" if self.where else self.message
        return f"{tag} {body}" + (f"\n           ↳ {self.hint}" if self.hint else "")


class DashboardError(Exception):
    """携带一组 Issue 的可预期失败。exit_code=1 表示数据问题，2 表示环境/用法问题。"""

    def __init__(self, issues, exit_code=1):
        self.issues = list(issues)
        self.exit_code = exit_code
        super().__init__(f"{len(self.issues)} 个问题")


@dataclass
class Report:
    """累积 issue，最后一次性抛出，避免用户修一个跑一次。"""

    issues: list = field(default_factory=list)

    def error(self, message, where="", hint=""):
        self.issues.append(Issue(message, where, "error", hint))

    def warn(self, message, where="", hint=""):
        self.issues.append(Issue(message, where, "warning", hint))

    @property
    def errors(self):
        return [i for i in self.issues if i.level == "error"]

    @property
    def warnings(self):
        return [i for i in self.issues if i.level == "warning"]

    def raise_if_errors(self):
        if self.errors:
            raise DashboardError(self.issues)


class CachedWorkbook:
    """公式缓存值的**惰性**读取器。

    openpyxl 读公式有个坑：data_only=False 拿到 '=A2*2' 字符串，data_only=True
    才拿得到 Excel 算好的缓存值，所以需要读两遍。但绝大多数模板里一个公式都没有，
    此前不管有没有公式都无条件把整个文件读第二遍——2000 行的表白花约 0.5 秒。

    这里改成只有真的撞上公式单元格时才打开第二份工作簿。
    也接受一个已经打开的工作簿对象（测试里会这么用）。
    """

    def __init__(self, source):
        self._path = None
        self._wb = None
        self._rows = {}
        if source is None:
            pass
        elif isinstance(source, (str, Path)):
            self._path = Path(source)
        else:
            self._wb = source

    @classmethod
    def wrap(cls, source):
        return source if isinstance(source, cls) else cls(source)

    @property
    def opened(self):
        """是否真的读了第二遍——性能测试拿它当断言。"""
        return self._wb is not None

    def _workbook(self):
        if self._wb is None and self._path is not None:
            self._wb = openpyxl.load_workbook(self._path, data_only=True)
        return self._wb

    def row_getter(self, sheet_title, offset, min_row=1):
        """返回一个零参可调用；**调用时**才会触发第二次读取。"""
        if self._path is None and self._wb is None:
            return None
        return lambda: self._row(sheet_title, offset, min_row)

    def _row(self, sheet_title, offset, min_row):
        key = (sheet_title, min_row)
        if key not in self._rows:
            wb = self._workbook()
            if wb is None or sheet_title not in wb.sheetnames:
                self._rows[key] = []
            else:
                self._rows[key] = list(wb[sheet_title].iter_rows(min_row=min_row))
        rows = self._rows[key]
        return rows[offset] if offset < len(rows) else None

    def close(self):
        if self._wb is not None and self._path is not None:
            self._wb.close()
        self._wb = None
        self._rows.clear()


# ---------------------------------------------------------------- 单元格解析

def parse_date(value, label):
    """把单元格值解析为 date；支持 datetime、日期串、Excel 序列号。"""
    if value is None or value == "":
        raise ValueError(f"{label}不能为空")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        d = from_excel(value)
        return d.date() if isinstance(d, datetime) else d
    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{label}不是有效日期（应为 yyyy-mm-dd，当前值：{text}）")


def parse_number(value, label, min_v=0, max_v=100):
    if value is None or value == "":
        raise ValueError(f"{label}不能为空")
    if isinstance(value, str):
        text = value.strip().rstrip("%").strip()
        if not re.fullmatch(r"\d+(\.\d+)?", text):
            raise ValueError(f"{label}应为 {min_v}～{max_v} 的数字，当前值：{value}")
        value = float(text)
    try:
        value = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{label}应为 {min_v}～{max_v} 的数字，当前值：{value}")
    if value < min_v or value > max_v:
        raise ValueError(f"{label}超出范围（{min_v}～{max_v}），当前值：{value}")
    return value


# ---------------------------------------------------------------- 工作表读取

def load_config(wb, report, cached_wb=None):
    """读取「项目配置」页。找不到时返回默认值。"""
    config = {
        "project_name": "未命名项目",
        "start": None,
        "end": None,
        "warn_days": 5,
        "workdays": "周一至周五",
    }
    cache = CachedWorkbook.wrap(cached_wb)
    for ws in wb.worksheets:
        if ws.title != "项目配置":
            continue
        for offset, row in enumerate(ws.iter_rows()):
            label = row[0].value
            cached_row = cache.row_getter(ws.title, offset)
            resolved, _ = resolve_values(row, cached_row, report, "项目配置", needed={1})
            value = resolved[1] if len(resolved) > 1 else None
            if not isinstance(label, str):
                continue
            label = label.strip()
            if label == "项目名称" and value:
                config["project_name"] = str(value).strip()
            elif label in ("项目开始日期", "项目结束日期"):
                key = "start" if label == "项目开始日期" else "end"
                try:
                    config[key] = parse_date(value, label)
                except ValueError as exc:
                    report.error(str(exc), "项目配置")
            elif label == "预警提前天数":
                try:
                    config["warn_days"] = max(0, int(float(str(value).strip())))
                except (TypeError, ValueError):
                    report.error("预警提前天数应为非负整数", "项目配置")
            elif label == "工作日" and value:
                config["workdays"] = str(value).strip()
    if config["start"] and config["end"] and config["start"] > config["end"]:
        report.error("项目开始日期晚于项目结束日期", "项目配置")
    return config


def resolve_values(raw_cells, cached_cells, report, where, needed=None):
    """把一行单元格解析成值；公式单元格取 Excel 缓存值。

    openpyxl 读公式有个坑：data_only=False 拿到的是 '=A2*2' 字符串，
    data_only=True 在文件没被 Excel 打开保存过时拿到 None。两边都读，
    只有「是公式且没有缓存值」才报错，并明确告诉用户怎么办。

    `cached_cells` 可以是一行单元格，也可以是一个零参可调用——后者到真正
    撞上公式时才求值，没有公式的表就不会被读第二遍（见 CachedWorkbook）。
    """
    values = []
    ok = True
    resolved_cache = _UNSET = object()
    for idx, cell in enumerate(raw_cells):
        value = cell.value
        if needed is not None and idx not in needed:
            values.append(value)
            continue
        if isinstance(value, str) and value.startswith("="):
            if resolved_cache is _UNSET:
                resolved_cache = (cached_cells() if callable(cached_cells)
                                  else cached_cells)
            row = resolved_cache
            cached = row[idx].value if row is not None and idx < len(row) else None
            if cached is None:
                report.error(
                    f"{cell.coordinate} 是公式（{value}）但没有计算结果缓存",
                    where,
                    hint="请用 Excel/WPS 打开该文件并保存一次，或把公式替换成数值",
                )
                value = None
                ok = False
            else:
                value = cached
        values.append(value)
    return values, ok


def load_options(wb):
    """读「选项字典」页，返回 {字段: {可选值...}}。没有该页就返回空字典。"""
    ws = next((w for w in wb.worksheets if w.title == "选项字典"), None)
    if ws is None:
        return {}
    options = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        field, value = row[0], row[1]
        if not isinstance(field, str) or value is None:
            continue
        field = field.strip()
        if field not in DICT_FIELDS:
            continue
        options.setdefault(field, set()).add(str(value).strip())
    return options


def check_options(tasks, options, report):
    """按「选项字典」校验枚举值。字典里没列的字段跳过；不匹配记 warning。"""
    for field, values in options.items():
        if not values:
            continue
        pick = DICT_FIELDS[field]
        for t in tasks:
            for got in pick(t):
                if got and got not in values:
                    report.warn(
                        f"{field}「{got}」不在「选项字典」页的候选里",
                        f"任务表第{t['row']}行（{t['id']}）",
                        hint="写错字会静默多出一个切片器分类；确属新增请先补进「选项字典」页",
                    )


def check_consistency(tasks, config, report):
    """跨字段一致性提示，全部记 warning，不阻塞出图。"""
    by_id = {t["id"]: t for t in tasks}
    for t in tasks:
        where = f"任务表第{t['row']}行（{t['id']}）"
        if t["status"] == "已完成" and t["progress"] < 100:
            report.warn(f"状态填了「已完成」，进度却是 {t['progress']:g}%", where,
                        hint="看板按「进度=100%」判定已完成，此行不会计入已完成卡")
        if t["progress"] >= 100 and t["status"] != "已完成":
            report.warn(f"进度已 100%，状态却还是「{t['status']}」", where)
        if config["start"] and t["end"] < config["start"]:
            report.warn(f"计划结束（{t['end']}）早于项目开始日期（{config['start']}）", where)
        if config["end"] and t["start"] > config["end"]:
            report.warn(f"计划开始（{t['start']}）晚于项目结束日期（{config['end']}）", where)
        for p in t["preds"]:
            pred = by_id.get(p)
            if pred and t["start"] < pred["start"]:
                # 后置任务比前置任务还早开工，排期与依赖关系真的矛盾了
                report.warn(
                    f"计划开始（{t['start']}）早于前置任务 {p} 的计划开始（{pred['start']}）",
                    where, hint="依赖倒挂：后置任务不可能先于前置任务开工")

    # 搭接（后置任务在前置任务完工前就启动）在快速跟进的排期里是常规做法，
    # 逐条报会淹没真正的问题，这里汇总成一条提示即可。
    overlaps = []
    for t in tasks:
        for p in t["preds"]:
            pred = by_id.get(p)
            if pred and pred["start"] <= t["start"] < pred["end"]:
                overlaps.append(f"{t['id']}←{p}（{(pred['end'] - t['start']).days} 天）")
    if overlaps:
        report.warn(
            f"{len(overlaps)} 处任务与其前置任务存在搭接：" + "、".join(overlaps),
            "任务表",
            hint="快速跟进排期下属正常；若本意是完成后再开始，请调整计划开始日期",
        )


def check_limits(tasks, timeline_start, timeline_end, report,
                 max_span=MAX_SPAN_DAYS, max_rows=MAX_ROWS, config=None):
    """输入规模护栏。超限直接报错，并指出是哪一行把范围撑开的。"""
    if len(tasks) > max_rows:
        report.error(
            f"任务行数 {len(tasks)} 超过上限 {max_rows}",
            "任务表",
            hint=f"当前渲染方式未做虚拟滚动；确需更多行请加 --max-rows {len(tasks)}",
        )

    span = (timeline_end - timeline_start).days + 1
    if span > max_span:
        # 时间轴上下界可能来自任务，也可能来自「项目配置」页；
        # 无配置日期时 build_payload 还会前后各留 7 天，所以不能靠等值反查。
        earliest = min(tasks, key=lambda t: t["start"], default=None)
        latest = max(tasks, key=lambda t: t["end"], default=None)
        lo = (f"第{earliest['row']}行（{earliest['id']}）计划开始 {earliest['start']}"
              if earliest else "（无任务）")
        hi = (f"第{latest['row']}行（{latest['id']}）计划结束 {latest['end']}"
              if latest else "（无任务）")
        if config and config.get("start") and earliest and config["start"] < earliest["start"]:
            lo = f"「项目配置」项目开始日期 {config['start']}"
        if config and config.get("end") and latest and config["end"] > latest["end"]:
            hi = f"「项目配置」项目结束日期 {config['end']}"
        report.error(
            f"时间轴跨度 {span} 天（{timeline_start} ~ {timeline_end}）超过上限 {max_span} 天",
            "任务表",
            hint=f"最可能是年份手误。下界来自 {lo}，上界来自 {hi}。"
                 f"确需更长周期请加 --max-span-days {span}",
        )


def find_task_sheet(wb):
    for ws in wb.worksheets:
        first_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
        first_row = [str(v).strip() if v is not None else "" for v in first_row]
        if "任务ID" in first_row and "阶段" in first_row and "任务" in first_row:
            return ws
    return None


def load_tasks(wb, report, cached_wb=None):
    """定位任务表并逐行校验，返回任务字典列表（16 字段模板）。"""
    task_sheet = find_task_sheet(wb)
    if task_sheet is None:
        report.error(
            "找不到任务表",
            hint="请确认某个工作表首行同时包含「任务ID」「阶段」「任务」这三个表头",
        )
        return []

    header_map = {}
    for idx, cell in enumerate(next(task_sheet.iter_rows(min_row=1, max_row=1))):
        name = str(cell.value).strip() if cell.value is not None else ""
        header_map[name] = idx
    missing = [h for h in HEADERS if h not in header_map]
    if missing:
        report.error(f"缺少必要列：{'、'.join(missing)}", "任务表")
        return []

    needed = {header_map[h] for h in HEADERS}
    cache = CachedWorkbook.wrap(cached_wb)

    tasks = []
    broken_ids = set()      # 本行自身有错、已被丢弃的任务ID，用于抑制下游的连锁误报
    filled_rows = 0         # 见到过的非空行数，用于区分「一行没填」和「填了但都不合法」
    for offset, cells in enumerate(task_sheet.iter_rows(min_row=2)):
        row_idx = offset + 2
        cached_cells = cache.row_getter(task_sheet.title, offset, min_row=2)
        resolved, resolved_ok = resolve_values(cells, cached_cells, report, f"任务表第{row_idx}行", needed)
        values = [resolved[header_map[h]] if header_map[h] < len(resolved) else None for h in HEADERS]
        raw = dict(zip(HEADERS, values))
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        filled_rows += 1
        if not resolved_ok:
            # 本行已因公式无缓存值报过错，再按字段校验只会刷出一堆「不能为空」；
            # 同时把它的 ID 记进 broken_ids，免得下游任务再被报「前置不存在」
            fallback_id = str(raw["任务ID"] or "").strip()
            if fallback_id:
                broken_ids.add(fallback_id)
            continue

        t = {"row": row_idx}
        where = f"任务表第{row_idx}行"
        try:
            task_id = str(raw["任务ID"] or "").strip()
            if not task_id:
                raise ValueError("任务ID不能为空")
            t["id"] = task_id
            where = f"任务表第{row_idx}行（{task_id}）"

            stage = str(raw["阶段"] or "").strip()
            if not stage:
                raise ValueError("阶段不能为空")
            t["stage"] = stage

            task_name = str(raw["任务"] or "").strip()
            if not task_name:
                raise ValueError("任务不能为空")
            t["task"] = task_name

            t["subtask"] = str(raw["子任务"] or "").strip()

            owner = str(raw["负责人"] or "").strip()
            owners = [o.strip() for o in re.split(r"[,，、;；]", owner) if o.strip()]
            if not owners:
                raise ValueError("负责人不能为空")
            t["owners"] = owners
            t["owner_text"] = "、".join(owners)

            t["dept"] = str(raw["协作部门"] or "").strip()

            t["start"] = parse_date(raw["计划开始"], "计划开始")
            t["end"] = parse_date(raw["计划结束"], "计划结束")
            if t["start"] > t["end"]:
                raise ValueError("计划开始日期晚于计划结束日期")

            t["progress"] = parse_number(raw["进度%"], "进度%", 0, 100)

            preds = str(raw["前置任务"] or "").strip()
            t["preds"] = [p.strip() for p in re.split(r"[,，;；、]", preds) if p.strip()]
            if task_id in t["preds"]:
                raise ValueError("前置任务不能包含自身")

            milestone = str(raw["里程碑"] or "否").strip()
            if milestone not in VALID_MILESTONE:
                raise ValueError(f"里程碑只能填 是/否，当前值：{raw['里程碑']}")
            t["milestone"] = milestone

            priority = str(raw["优先级"] or "中").strip()
            if priority not in VALID_PRIORITY:
                raise ValueError(f"优先级只能填 高/中/低，当前值：{raw['优先级']}")
            t["priority"] = priority

            status = str(raw["状态"] or "未开始").strip()
            if status not in VALID_STATUS:
                raise ValueError(f"状态只能填 未开始/进行中/待评审/已完成/阻塞，当前值：{raw['状态']}")
            t["status"] = status

            risk = str(raw["风险等级"] or "中").strip()
            if risk not in VALID_RISK:
                raise ValueError(f"风险等级只能填 高/中/低，当前值：{raw['风险等级']}")
            t["risk"] = risk

            t["deliverable"] = str(raw["交付物"] or "").strip()
            t["note"] = str(raw["备注"] or "").strip()
            tasks.append(t)
        except ValueError as exc:
            report.error(str(exc), where)
            if t.get("id"):
                broken_ids.add(t["id"])

    if filled_rows == 0:
        # 填了但都不合法时不报这条——逐行错误已经说明问题了
        report.error("任务页没有数据：请至少填写一行任务", "任务表")

    check_graph(tasks, report, broken_ids)
    return tasks


def check_graph(tasks, report, broken_ids=frozenset()):
    """任务ID 唯一性、前置任务存在性、依赖无环——三类一次全查完，不再逐个中断。"""
    seen = set()
    for t in tasks:
        if t["id"] in seen:
            report.error(f"任务ID「{t['id']}」重复", f"任务表第{t['row']}行")
        seen.add(t["id"])

    by_id = {t["id"]: t for t in tasks}
    for t in tasks:
        for p in t["preds"]:
            if p in by_id:
                continue
            # 前置任务本身那一行已经报过错了，这里不再重复制造连锁误报
            if p in broken_ids:
                continue
            report.error(f"前置任务「{p}」不存在", f"任务表第{t['row']}行（{t['id']}）")

    cycle = detect_cycle(tasks)
    if cycle:
        report.error(f"前置任务存在循环依赖：{' → '.join(cycle)}", "任务表")


def detect_cycle(tasks):
    """检测前置任务环，返回环路径（含首尾重复）或 None。

    用显式栈迭代，不用递归：逆序填写的长依赖链会让递归 DFS 一次探到底，
    实测 1200 级就会抛 RecursionError，且是未捕获的原始 traceback。
    """
    WHITE, GREY, BLACK = 0, 1, 2
    by_id = {t["id"]: t for t in tasks}
    mark = {tid: WHITE for tid in by_id}

    for root in by_id:
        if mark[root] != WHITE:
            continue
        mark[root] = GREY
        path = [root]
        stack = [(root, iter(by_id[root]["preds"]))]
        while stack:
            node, pending = stack[-1]
            descended = False
            for pred in pending:
                if pred not in by_id:
                    continue
                if mark[pred] == GREY:                 # 回边 → 成环
                    return path[path.index(pred):] + [pred]
                if mark[pred] == WHITE:
                    mark[pred] = GREY
                    path.append(pred)
                    stack.append((pred, iter(by_id[pred]["preds"])))
                    descended = True
                    break
            if not descended:
                mark[node] = BLACK
                stack.pop()
                path.pop()
    return None


# ---------------------------------------------------------------- 关键路径

def topological_order(tasks):
    """Kahn 拓扑排序，前置在前。调用前已确认无环。"""
    by_id = {t["id"]: t for t in tasks}
    preds = {t["id"]: [p for p in t["preds"] if p in by_id] for t in tasks}
    indeg = {tid: len(ps) for tid, ps in preds.items()}
    succs = {tid: [] for tid in by_id}
    for tid, ps in preds.items():
        for p in ps:
            succs[p].append(tid)

    queue = [tid for tid, d in indeg.items() if d == 0]
    order = []
    while queue:
        tid = queue.pop(0)
        order.append(tid)
        for s in succs[tid]:
            indeg[s] -= 1
            if indeg[s] == 0:
                queue.append(s)
    return order, succs


def compute_schedule(tasks):
    """CPM 正推 / 逆推，返回 {任务ID: {es, ef, ls, lf, tf, critical}}（均为序数日）。

    口径说明：按**日历日**计算。「项目配置」页的「工作日」字段目前只作展示，
    不参与任何推算——改成按工作日推算会连带影响工期、条形宽度与预警判定。

    - 无前置的任务以其计划开始作为最早开始（否则整网没有时间锚点）
    - ES = max(前置 EF) + 1，EF = ES + 工期 - 1
    - 自最晚 EF 逆推 LF / LS，总浮动时差 TF = LS - ES，TF == 0 即关键路径
    """
    if not tasks:
        return {}
    by_id = {t["id"]: t for t in tasks}
    order, succs = topological_order(tasks)
    if len(order) != len(by_id):        # 有环，理论上已被 check_graph 拦下
        return {}

    dur = {t["id"]: (t["end"] - t["start"]).days + 1 for t in tasks}
    es, ef = {}, {}
    for tid in order:
        preds = [p for p in by_id[tid]["preds"] if p in by_id]
        es[tid] = max((ef[p] for p in preds), default=None)
        es[tid] = es[tid] + 1 if es[tid] is not None else by_id[tid]["start"].toordinal()
        ef[tid] = es[tid] + dur[tid] - 1

    finish = max(ef.values())
    ls, lf = {}, {}
    for tid in reversed(order):
        lf[tid] = min((ls[s] for s in succs[tid]), default=finish + 1) - 1 if succs[tid] else finish
        ls[tid] = lf[tid] - dur[tid] + 1

    return {tid: {"es": es[tid], "ef": ef[tid], "ls": ls[tid], "lf": lf[tid],
                  "tf": ls[tid] - es[tid], "critical": ls[tid] - es[tid] == 0}
            for tid in by_id}


# ---------------------------------------------------------------- 数据装配

def build_payload(tasks, config, today):
    """生成内嵌到 HTML 的标准 JSON 数据。"""
    sched = compute_schedule(tasks)
    ord_to_iso = lambda n: date.fromordinal(n).isoformat()
    payload_tasks = []
    for t in tasks:
        display_name = t["task"] if not t["subtask"] else f"{t['task']} · {t['subtask']}"
        payload_tasks.append(
            {
                "id": t["id"],
                "stage": t["stage"],
                "task": t["task"],
                "subtask": t["subtask"],
                "name": display_name,
                "owners": t["owners"],
                "ownerText": t["owner_text"],
                "dept": t["dept"],
                "start": t["start"].isoformat(),
                "end": t["end"].isoformat(),
                "duration": (t["end"] - t["start"]).days + 1,
                "progress": round(t["progress"], 1),
                "preds": t["preds"],
                "milestone": t["milestone"] == "是",
                "priority": t["priority"],
                "status": t["status"],
                "risk": t["risk"],
                "deliverable": t["deliverable"],
                "note": t["note"],
            }
        )
        cpm = sched.get(t["id"])
        if cpm:
            payload_tasks[-1].update({
                "es": ord_to_iso(cpm["es"]), "ef": ord_to_iso(cpm["ef"]),
                "ls": ord_to_iso(cpm["ls"]), "lf": ord_to_iso(cpm["lf"]),
                "tf": cpm["tf"], "critical": cpm["critical"],
            })

    starts = [t["start"] for t in tasks]
    ends = [t["end"] for t in tasks]
    if config["start"] and config["end"]:
        timeline_start = min([config["start"]] + starts)
        timeline_end = max([config["end"]] + ends)
    else:
        timeline_start = min(starts) - timedelta(days=7)
        timeline_end = max(ends) + timedelta(days=7)

    crit_ids = [t["id"] for t in payload_tasks if t.get("critical")]
    crit_span = 0
    if crit_ids:
        crit_span = (max(sched[i]["ef"] for i in crit_ids)
                     - min(sched[i]["es"] for i in crit_ids) + 1)
    return {
        "project": config["project_name"],
        "warnDays": config["warn_days"],
        "criticalCount": len(crit_ids),
        "criticalDuration": crit_span,
        "workdays": config["workdays"],
        "timelineStart": timeline_start.isoformat(),
        "timelineEnd": timeline_end.isoformat(),
        "buildDate": today.isoformat(),
        "tasks": payload_tasks,
    }


# ---------------------------------------------------------------- 渲染

def load_assets(asset_dir=ASSET_DIR):
    missing = [n for n in ASSET_NAMES if not (asset_dir / n).is_file()]
    if missing:
        raise DashboardError([Issue(
            f"缺少前端资源文件：{'、'.join(missing)}",
            str(asset_dir),
            hint="assets/ 需与 build_dashboard.py 放在同一目录，请确认没有被移动或漏拷贝",
        )], exit_code=2)
    return {n: (asset_dir / n).read_text(encoding="utf-8") for n in ASSET_NAMES}


def data_script(payload):
    """把 payload 序列化成 JS 赋值语句；转义 </ 防止提前闭合 script 标签。"""
    return "const DATA = " + json.dumps(payload, ensure_ascii=False).replace("</", "<\\/") + ";\n"


def render_html(assets, payload):
    """把 CSS / DATA / JS 内联进模板，产出单文件自包含 HTML。"""
    css, js = assets["dashboard.css"], assets["dashboard.js"]
    html = assets["template.html"]

    if "</style>" in css:
        raise DashboardError([Issue("dashboard.css 含有 </style> 字面量，内联后会提前闭合样式块", "assets")], exit_code=2)
    if "</script>" in js:
        raise DashboardError([Issue("dashboard.js 含有 </script> 字面量，内联后会提前闭合脚本块", "assets")], exit_code=2)
    for tag in (TAG_CSS, TAG_DATA, TAG_JS):
        if tag not in html:
            raise DashboardError([Issue(f"template.html 缺少内联锚点：{tag}", "assets")], exit_code=2)

    html = html.replace(TAG_CSS, f"<style>\n{css}</style>")
    html = html.replace(TAG_DATA, f"<script>\n{data_script(payload)}</script>")
    html = html.replace(TAG_JS, f"<script>\n{js}</script>")
    return html


EXTERNAL_REF = re.compile(r'(?:src|href)\s*=\s*["\'](?!#|data:)([^"\']+)')


def external_refs(html):
    """返回产物里残留的外部引用，用于自检「离线自包含」这一硬约束。"""
    return EXTERNAL_REF.findall(html)


# ---------------------------------------------------------------- 命令行

def build(input_path, asset_dir=ASSET_DIR, today=None, strict=False,
          max_span=MAX_SPAN_DAYS, max_rows=MAX_ROWS):
    """读表 → 校验 → 装配 payload。抛 DashboardError 表示数据有问题。"""
    if not input_path.exists():
        raise DashboardError([Issue(f"找不到输入文件：{input_path}")], exit_code=2)
    try:
        wb = openpyxl.load_workbook(input_path, data_only=False)
    except Exception as exc:
        raise DashboardError([Issue(f"无法打开 Excel 文件：{exc}", str(input_path))], exit_code=2)

    # 带缓存值的第二份工作簿是惰性的：没有公式单元格就永远不会被打开
    cache = CachedWorkbook(input_path)
    report = Report()
    try:
        config = load_config(wb, report, cache)
        tasks = load_tasks(wb, report, cache)
        if tasks:
            check_options(tasks, load_options(wb), report)
            check_consistency(tasks, config, report)
    finally:
        cache.close()
        wb.close()
    report.raise_if_errors()

    payload = build_payload(tasks, config, today or date.today())
    check_limits(tasks,
                 date.fromisoformat(payload["timelineStart"]),
                 date.fromisoformat(payload["timelineEnd"]),
                 report, max_span, max_rows, config)

    if strict:
        for issue in report.warnings:
            issue.level = "error"
            issue.hint = (issue.hint + "（--strict 下提示按错误处理）").lstrip("（")
    report.raise_if_errors()

    payload["issues"] = [
        {"level": i.level, "where": i.where, "message": i.message, "hint": i.hint}
        for i in report.issues
    ]
    return payload, report


def main(argv=None):
    parser = argparse.ArgumentParser(description="从 Excel 模板生成离线项目看板")
    parser.add_argument("--input", default="家电产品项目_任务数据.xlsx", help="16 字段 Excel 模板路径")
    parser.add_argument("--output", default="项目看板.html", help="输出的 HTML 路径")
    parser.add_argument("--emit-json", action="store_true", help="只输出标准 JSON 数据，不生成 HTML")
    parser.add_argument("--dev", action="store_true",
                        help="额外写出 assets/data.js，便于直接打开 assets/template.html 调试")
    parser.add_argument("--assets", default=str(ASSET_DIR), help="前端资源目录")
    parser.add_argument("--strict", action="store_true", help="把提示（warning）也当作错误处理")
    parser.add_argument("--max-span-days", type=int, default=MAX_SPAN_DAYS,
                        help=f"时间轴跨度上限，默认 {MAX_SPAN_DAYS} 天")
    parser.add_argument("--max-rows", type=int, default=MAX_ROWS,
                        help=f"任务行数上限，默认 {MAX_ROWS} 行")
    parser.add_argument("--traceback", action="store_true", help="出错时打印完整堆栈")
    args = parser.parse_args(argv)

    asset_dir = Path(args.assets)
    try:
        payload, report = build(Path(args.input), asset_dir, strict=args.strict,
                                max_span=args.max_span_days, max_rows=args.max_rows)

        if args.emit_json:
            print(json.dumps(payload, ensure_ascii=False, indent=2))
            return 0

        assets = load_assets(asset_dir)
        html = render_html(assets, payload)

        leftovers = external_refs(html)
        if leftovers:
            raise DashboardError([Issue(
                f"产物残留外部引用：{'、'.join(leftovers)}",
                str(args.output),
                hint="离线看板必须自包含，请检查 assets/ 里是否引入了 CDN 或本地文件",
            )], exit_code=2)

        Path(args.output).write_text(html, encoding="utf-8")
        if args.dev:
            (asset_dir / "data.js").write_text(data_script(payload), encoding="utf-8")

        for issue in report.warnings:
            print(issue.format(), file=sys.stderr)
        warn_note = f"，{len(report.warnings)} 条提示（看板顶栏可查看）" if report.warnings else ""
        print(f"已生成 {args.output}（{len(payload['tasks'])} 个任务{warn_note}）"
              + (f"，另已写出 {asset_dir / 'data.js'}" if args.dev else ""))
        return 0

    except DashboardError as exc:
        for issue in sorted(exc.issues, key=lambda i: i.level != "error"):
            print(issue.format(), file=sys.stderr)
        n = sum(1 for i in exc.issues if i.level == "error")
        tail = "修正后重新运行即可。" if exc.exit_code == 1 else "这是环境或用法问题，不是表格数据的问题。"
        print(f"\n共 {n} 个错误，已中止生成。{tail}", file=sys.stderr)
        return exc.exit_code
    except OSError as exc:
        print(f"[读写失败] {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # 兜底：不让原始 traceback 直接弹到 .command 窗口
        print(f"[意外错误] {type(exc).__name__}: {exc}", file=sys.stderr)
        if args.traceback:
            _traceback.print_exc()
        else:
            print("加 --traceback 可查看完整堆栈。", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
