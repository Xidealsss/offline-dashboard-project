# -*- coding: utf-8 -*-
"""阶段 3：输入护栏、公式单元格、选项字典校验、一致性提示、--strict。"""

import tempfile
import unittest
from datetime import date
from pathlib import Path

import openpyxl

from helpers import bd, make_workbook, run_cli, task


def build_from(rows, **kw):
    """写临时 xlsx 后走完整 build()，返回 (payload, report) 或抛 DashboardError。"""
    with tempfile.TemporaryDirectory() as d:
        p = make_workbook(Path(d) / "t.xlsx", rows, **{k: v for k, v in kw.items()
                                                       if k in ("config", "options")})
        return bd.build(p, **{k: v for k, v in kw.items()
                              if k in ("strict", "max_span", "max_rows", "today")})


def issues_of(cm):
    return [(i.level, i.message) for i in cm.exception.issues]


class TestSpanGuard(unittest.TestCase):
    def test_year_typo_is_caught_before_the_browser_chokes(self):
        """把 2026 打成 2226 会撑出 7 万天时间轴、14 万个表头节点。"""
        with self.assertRaises(bd.DashboardError) as cm:
            build_from([task(任务ID="T1", 计划开始=date(2026, 1, 1), 计划结束=date(2226, 12, 31))],
                       config=None)
        text = " ".join(m for _, m in issues_of(cm))
        self.assertIn("时间轴跨度", text)
        self.assertIn("超过上限", text)

    def test_error_points_at_the_offending_row(self):
        with self.assertRaises(bd.DashboardError) as cm:
            build_from([task(任务ID="T1"),
                        task(任务ID="T2", 计划开始=date(2026, 1, 1), 计划结束=date(2226, 12, 31))],
                       config=None)
        hint = " ".join(i.hint for i in cm.exception.issues)
        self.assertIn("T2", hint)
        self.assertIn("2226-12-31", hint)

    def test_limit_is_overridable(self):
        payload, _ = build_from(
            [task(任务ID="T1", 计划开始=date(2026, 1, 1), 计划结束=date(2031, 12, 31))],
            config=None, max_span=99999)
        self.assertEqual(payload["timelineEnd"], "2032-01-07")

    def test_normal_project_passes(self):
        payload, _ = build_from([task()], config=None)
        self.assertTrue(payload["tasks"])


class TestRowCountGuard(unittest.TestCase):
    def test_over_limit_is_rejected(self):
        rows = [task(任务ID=f"T{i}") for i in range(30)]
        with self.assertRaises(bd.DashboardError) as cm:
            build_from(rows, config=None, max_rows=10)
        self.assertIn("任务行数 30 超过上限 10", " ".join(m for _, m in issues_of(cm)))

    def test_hint_tells_you_the_flag_to_use(self):
        with self.assertRaises(bd.DashboardError) as cm:
            build_from([task(任务ID=f"T{i}") for i in range(5)], config=None, max_rows=2)
        self.assertIn("--max-rows 5", " ".join(i.hint for i in cm.exception.issues))


class TestFormulaCells(unittest.TestCase):
    def write_with_formula(self, d, cell, formula):
        p = make_workbook(Path(d) / "f.xlsx", [task(任务ID="T1"), task(任务ID="T2", 前置任务="T1")],
                          config=None)
        wb = openpyxl.load_workbook(p)
        wb["任务"][cell] = formula
        wb.save(p)
        return p

    def test_formula_without_cached_value_gives_actionable_error(self):
        with tempfile.TemporaryDirectory() as d:
            p = self.write_with_formula(d, "I2", "=50*2")
            with self.assertRaises(bd.DashboardError) as cm:
                bd.build(p)
        joined = " ".join(m for _, m in issues_of(cm))
        self.assertIn("是公式", joined)
        hints = " ".join(i.hint for i in cm.exception.issues)
        self.assertIn("保存一次", hints)

    def test_broken_formula_row_does_not_cascade(self):
        """T1 因公式无缓存被丢弃后，不该再为 T2 报「前置 T1 不存在」。"""
        with tempfile.TemporaryDirectory() as d:
            p = self.write_with_formula(d, "I2", "=50*2")
            with self.assertRaises(bd.DashboardError) as cm:
                bd.build(p)
        msgs = [m for lvl, m in issues_of(cm) if lvl == "error"]
        self.assertEqual(len(msgs), 1, f"应只报公式那一条，实际：{msgs}")

    def test_cached_value_is_used_when_present(self):
        """模拟 Excel 保存过的文件：公式带缓存值时应当直接采用。"""
        with tempfile.TemporaryDirectory() as d:
            p = make_workbook(Path(d) / "c.xlsx", [task(任务ID="T1")], config=None)
            wb = openpyxl.load_workbook(p)
            wb["任务"]["I2"] = "=50*2"
            wb.save(p)
            # openpyxl 不算公式，手工塞一份缓存值模拟 Excel 的行为
            cached = openpyxl.load_workbook(p, data_only=True)
            cached["任务"]["I2"] = 100
            cached.save(Path(d) / "cached.xlsx")

            report = bd.Report()
            tasks = bd.load_tasks(openpyxl.load_workbook(p),
                                  report,
                                  openpyxl.load_workbook(Path(d) / "cached.xlsx", data_only=True))
        self.assertEqual([i.message for i in report.errors], [])
        self.assertEqual(tasks[0]["progress"], 100.0)


class TestOptionDictionary(unittest.TestCase):
    OPTIONS = {"阶段": ["阶段甲", "阶段乙"], "负责人": ["张三", "李四"], "风险等级": ["高", "中", "低"]}

    def test_value_outside_dictionary_is_a_warning_not_an_error(self):
        payload, report = build_from([task(负责人="李四四")], config=None, options=self.OPTIONS)
        self.assertEqual(report.errors, [], "字典不匹配不应阻塞出图")
        self.assertTrue(any("不在「选项字典」" in i.message for i in report.warnings))

    def test_values_inside_dictionary_are_silent(self):
        _, report = build_from([task(阶段="阶段乙", 负责人="张三")], config=None, options=self.OPTIONS)
        self.assertEqual([i.message for i in report.warnings], [])

    def test_trailing_whitespace_is_normalised_before_checking(self):
        _, report = build_from([task(阶段="阶段甲  ")], config=None, options=self.OPTIONS)
        self.assertEqual([i.message for i in report.warnings], [])

    def test_fields_absent_from_dictionary_are_skipped(self):
        _, report = build_from([task(协作部门="没在字典里的部门")], config=None, options=self.OPTIONS)
        self.assertEqual([i.message for i in report.warnings], [])

    def test_no_dictionary_sheet_means_no_checks(self):
        _, report = build_from([task(负责人="随便谁")], config=None, options=None)
        self.assertEqual([i.message for i in report.warnings], [])

    def test_each_owner_in_a_multi_owner_cell_is_checked(self):
        _, report = build_from([task(负责人="张三、王五")], config=None, options=self.OPTIONS)
        self.assertTrue(any("王五" in i.message for i in report.warnings))


class TestConsistency(unittest.TestCase):
    CFG = {"项目名称": "P", "项目开始日期": date(2026, 1, 1),
           "项目结束日期": date(2026, 12, 31), "预警提前天数": 5}

    def warns(self, rows, config=None):
        _, report = build_from(rows, config=config or self.CFG)
        return " | ".join(f"{i.message} {i.hint}" for i in report.warnings)

    def test_status_done_but_progress_incomplete(self):
        self.assertIn("状态填了「已完成」",
                      self.warns([task(状态="已完成", **{"进度%": 40})]))

    def test_progress_complete_but_status_not_done(self):
        self.assertIn("进度已 100%",
                      self.warns([task(状态="进行中", **{"进度%": 100})]))

    def test_task_entirely_before_project_start(self):
        self.assertIn("早于项目开始日期",
                      self.warns([task(计划开始=date(2025, 1, 1), 计划结束=date(2025, 2, 1))]))

    def test_task_starting_after_project_end(self):
        self.assertIn("晚于项目结束日期",
                      self.warns([task(计划开始=date(2027, 1, 1), 计划结束=date(2027, 2, 1))]))

    def test_true_dependency_inversion_is_flagged(self):
        """后置任务比前置任务还早开工——排期真的矛盾。"""
        text = self.warns([
            task(任务ID="T1", 计划开始=date(2026, 6, 1), 计划结束=date(2026, 6, 20)),
            task(任务ID="T2", 前置任务="T1", 计划开始=date(2026, 5, 1), 计划结束=date(2026, 7, 1)),
        ])
        self.assertIn("依赖倒挂", text)
        self.assertIn("早于前置任务 T1 的计划开始", text)

    def test_overlap_is_summarised_into_a_single_line(self):
        """搭接是快速跟进的常规做法，逐条报会淹没真正的问题。"""
        rows = [task(任务ID="T0", 计划开始=date(2026, 3, 1), 计划结束=date(2026, 4, 1))]
        for i in range(1, 6):
            rows.append(task(任务ID=f"T{i}", 前置任务="T0",
                             计划开始=date(2026, 3, 15), 计划结束=date(2026, 5, 1)))
        _, report = build_from(rows, config=self.CFG)
        overlap = [i for i in report.warnings if "搭接" in i.message]
        self.assertEqual(len(overlap), 1, "5 处搭接应汇总成 1 条提示")
        self.assertIn("5 处", overlap[0].message)

    def test_clean_plan_produces_no_warnings(self):
        _, report = build_from([
            task(任务ID="T1", 计划开始=date(2026, 2, 1), 计划结束=date(2026, 2, 10)),
            task(任务ID="T2", 前置任务="T1", 计划开始=date(2026, 2, 11), 计划结束=date(2026, 2, 20)),
        ], config=self.CFG)
        self.assertEqual([i.message for i in report.warnings], [])


class TestStrictMode(unittest.TestCase):
    def test_warnings_stay_warnings_by_default(self):
        payload, report = build_from([task(状态="已完成", **{"进度%": 40})], config=None)
        self.assertTrue(report.warnings)
        self.assertEqual(report.errors, [])
        self.assertTrue(payload["tasks"])

    def test_strict_promotes_warnings_to_errors(self):
        with self.assertRaises(bd.DashboardError) as cm:
            build_from([task(状态="已完成", **{"进度%": 40})], config=None, strict=True)
        self.assertEqual(cm.exception.exit_code, 1)
        self.assertTrue(all(i.level == "error" for i in cm.exception.issues))

    def test_strict_via_cli(self):
        with tempfile.TemporaryDirectory() as d:
            p = make_workbook(Path(d) / "w.xlsx", [task(状态="已完成", **{"进度%": 40})], config=None)
            ok, _, _ = run_cli("--input", p, "--output", Path(d) / "a.html")
            strict, _, err = run_cli("--input", p, "--output", Path(d) / "b.html", "--strict")
        self.assertEqual(ok, 0)
        self.assertEqual(strict, 1)
        self.assertIn("[校验失败]", err)


class TestIssuesReachThePayload(unittest.TestCase):
    def test_warnings_are_embedded_for_the_dashboard_to_show(self):
        payload, report = build_from([task(状态="已完成", **{"进度%": 40})], config=None)
        self.assertTrue(payload["issues"])
        self.assertEqual(len(payload["issues"]), len(report.issues))
        self.assertEqual(payload["issues"][0]["level"], "warning")

    def test_clean_project_embeds_an_empty_list(self):
        payload, _ = build_from([task()], config=None)
        self.assertEqual(payload["issues"], [])


class TestCliMessaging(unittest.TestCase):
    def test_errors_are_printed_before_warnings(self):
        with tempfile.TemporaryDirectory() as d:
            p = make_workbook(Path(d) / "m.xlsx", [
                task(任务ID="T1", 状态="已完成", **{"进度%": 40}),          # warning
                task(任务ID="T2", 计划开始=date(2026, 1, 1), 计划结束=date(2226, 1, 1)),  # error
            ], config=None)
            _, _, err = run_cli("--input", p, "--output", Path(d) / "o.html")
        first_error = err.index("[校验失败]")
        first_warning = err.index("[提示]")
        self.assertLess(first_error, first_warning, "错误必须排在提示前面")

    def test_success_line_mentions_warning_count(self):
        with tempfile.TemporaryDirectory() as d:
            p = make_workbook(Path(d) / "w.xlsx", [task(状态="已完成", **{"进度%": 40})], config=None)
            code, out, _ = run_cli("--input", p, "--output", Path(d) / "o.html")
        self.assertEqual(code, 0)
        self.assertIn("条提示", out)


if __name__ == "__main__":
    unittest.main()
