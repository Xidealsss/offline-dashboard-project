# -*- coding: utf-8 -*-
"""测试公共设施：动态加载被测脚本、按 16 字段模板造 Excel、跑命令行。

只依赖标准库 + openpyxl（生成看板本来就需要它），不引入任何测试框架依赖。
"""

import importlib.util
import subprocess
import sys
from datetime import date
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPT = PROJECT_ROOT / "build_dashboard.py"
REAL_XLSX = PROJECT_ROOT / "家电产品项目_任务数据.xlsx"


def load_build():
    """把 build_dashboard.py 当模块导入，便于直接单测内部函数。"""
    spec = importlib.util.spec_from_file_location("build_dashboard", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


bd = load_build()
HEADERS = bd.HEADERS

DEFAULT_CONFIG = {
    "项目名称": "单测项目",
    "项目开始日期": date(2026, 1, 1),
    "项目结束日期": date(2026, 6, 30),
    "预警提前天数": 5,
    "工作日": "周一至周五",
}

_TASK_DEFAULTS = {
    "任务ID": "T001",
    "阶段": "阶段甲",
    "任务": "示例任务",
    "子任务": "",
    "负责人": "张三",
    "协作部门": "研发",
    "计划开始": date(2026, 1, 5),
    "计划结束": date(2026, 1, 9),
    "进度%": 0,
    "前置任务": "",
    "里程碑": "否",
    "优先级": "中",
    "状态": "未开始",
    "风险等级": "中",
    "交付物": "",
    "备注": "",
}


def task(**overrides):
    """造一行任务；只写关心的字段，其余取默认值。"""
    row = dict(_TASK_DEFAULTS)
    unknown = set(overrides) - set(_TASK_DEFAULTS)
    if unknown:
        raise KeyError(f"未知字段：{unknown}")
    row.update(overrides)
    return row


def chain(n, reverse=False, span=2):
    """造一条 T0 <- T1 <- ... <- T(n-1) 的前置链，用于压环检测的递归深度。"""
    rows = []
    for i in range(n):
        rows.append(task(
            任务ID=f"T{i}",
            前置任务=f"T{i - 1}" if i else "",
            计划开始=date(2026, 1, 1),
            计划结束=date(2026, 1, span),
        ))
    return rows[::-1] if reverse else rows


def make_workbook(path, rows=None, config=DEFAULT_CONFIG, options=None, headers=None):
    """按 16 字段模板写一个 xlsx。

    rows      任务行（task() 产物）列表；None 表示只有表头
    config    「项目配置」页键值；None 表示不建该页
    options   「选项字典」页 {字段: [可选值...]}；None 表示不建该页
    headers   覆盖表头，用于测试缺列/改名
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务"
    ws.append(list(headers if headers is not None else HEADERS))
    for row in rows or []:
        ws.append([row.get(h, "") for h in HEADERS])

    if config is not None:
        cfg = wb.create_sheet("项目配置")
        cfg.append(["项目配置", ""])
        for k, v in config.items():
            cfg.append([k, v])

    if options is not None:
        opt = wb.create_sheet("选项字典")
        opt.append(["选项字典（飞书单选字段选项）", ""])
        opt.append(["字段", "选项值"])
        for field, values in options.items():
            for v in values:
                opt.append([field, v])

    wb.save(path)
    return Path(path)


def collect(rows=None, **kw):
    """跑一遍解析+校验，返回 (tasks, report)，不抛异常，便于断言 issue 列表。"""
    import tempfile
    with tempfile.TemporaryDirectory() as d:
        p = make_workbook(Path(d) / "t.xlsx", rows, **kw)
        wb = openpyxl.load_workbook(p, data_only=False)
        report = bd.Report()
        bd.load_config(wb, report)
        tasks = bd.load_tasks(wb, report)
        return tasks, report


def messages(report, level=None):
    return [i.message for i in report.issues if level is None or i.level == level]


def run_cli(*args, cwd=None):
    """跑一次命令行，返回 (returncode, stdout, stderr)。"""
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), *map(str, args)],
        capture_output=True, text=True, cwd=str(cwd or PROJECT_ROOT),
    )
    return proc.returncode, proc.stdout, proc.stderr
